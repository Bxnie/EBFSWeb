#!/usr/bin/env python3
"""
Generates ticket-data.json from the Bristol Ticket Counts spreadsheet.
Reads the 'Live' and 'Club' sheets (summary sheets) and outputs the same
JSON format as ticket-data.json.

Usage:
    python3 generate_ticket_data.py "22.05.26 Bristol Ticket Counts.xlsx"
    python3 generate_ticket_data.py  # uses the first .xlsx found in current dir
"""

import json
import sys
import os
import glob
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print("Error: openpyxl is required. Run: pip install openpyxl")
    sys.exit(1)

MONTH_NAMES = {
    1: "JANUARY", 2: "FEBRUARY", 3: "MARCH", 4: "APRIL",
    5: "MAY", 6: "JUNE", 7: "JULY", 8: "AUGUST",
    9: "SEPTEMBER", 10: "OCTOBER", 11: "NOVEMBER", 12: "DECEMBER",
}

# Column indices (0-based) in the Live/Club summary sheets
COL_DATE = 0
COL_EVENT = 3
COL_PROMOTER = 4
COL_CAPACITY = 5
COL_TOTAL_SALES = 7  # "Total Sales" — the current live count


def parse_capacity(value):
    """Convert capacity cell value (e.g. '1,300', '3000', 1300) to int."""
    if value is None:
        return None
    return int(str(value).replace(",", "").strip())


def extract_shows(ws, sheet_type):
    """Extract all show rows from a summary sheet (Live or Club)."""
    shows = []
    for row in ws.iter_rows(min_row=1, values_only=True):
        event_date = row[COL_DATE]
        if not isinstance(event_date, datetime):
            continue
        capacity_raw = row[COL_CAPACITY]
        if capacity_raw is None or str(capacity_raw).strip().lower() in ("", "capacity"):
            continue
        capacity = parse_capacity(capacity_raw)
        if not capacity:
            continue
        event_name = row[COL_EVENT]
        if not event_name or not isinstance(event_name, str):
            continue
        event_name = event_name.strip()
        promoter = row[COL_PROMOTER]
        promoter = promoter.strip() if isinstance(promoter, str) else str(promoter or "")
        total_sales_raw = row[COL_TOTAL_SALES]
        try:
            sold = int(total_sales_raw) if total_sales_raw not in (None, "", "#REF!") else 0
        except (ValueError, TypeError):
            sold = 0
        pct = round(sold / capacity, 10) if capacity else 0.0
        shows.append({
            "date": event_date.strftime("%Y-%m-%d"),
            "day": event_date.day,
            "month": event_date.month,
            "monthName": MONTH_NAMES[event_date.month],
            "year": event_date.year,
            "name": event_name,
            "promoter": promoter,
            "capacity": capacity,
            "sold": sold,
            "pct": pct,
            "type": sheet_type,
        })
    return shows


def main():
    if len(sys.argv) > 1:
        xlsx_path = sys.argv[1]
    else:
        matches = glob.glob("*.xlsx")
        if not matches:
            print("Error: no .xlsx file found in current directory.")
            sys.exit(1)
        xlsx_path = matches[0]
        print(f"Using spreadsheet: {xlsx_path}")

    if not os.path.exists(xlsx_path):
        print(f"Error: file not found: {xlsx_path}")
        sys.exit(1)

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    all_shows = []
    seen_dates = set()

    for sheet_name in ("Live", "Club"):
        if sheet_name not in wb.sheetnames:
            print(f"Warning: sheet '{sheet_name}' not found, skipping.")
            continue
        ws = wb[sheet_name]
        sheet_type = sheet_name.lower()
        for show in extract_shows(ws, sheet_type):
            key = (show["date"], show["name"])
            if key not in seen_dates:
                seen_dates.add(key)
                all_shows.append(show)

    all_shows.sort(key=lambda s: s["date"])

    out_path = "ticket-data.json"
    with open(out_path, "w") as f:
        json.dump(all_shows, f, indent=2)
        f.write("\n")

    print(f"Written {len(all_shows)} shows to {out_path}")


if __name__ == "__main__":
    main()
